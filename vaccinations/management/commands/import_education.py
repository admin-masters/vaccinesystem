from __future__ import annotations
from pathlib import Path
from typing import Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q

from openpyxl import load_workbook

from vaccinations.models import (
    ScheduleVersion, Vaccine, VaccineEducationPatient, VaccineEducationDoctor, VideoPlatform
)

REQUIRED_PE = {"video_url", "language"}
REQUIRED_DE = {"video_url"}

def _norm(s: Optional[str]) -> str:
    return (s or "").strip()

def _yn(s: Optional[str]) -> bool:
    return (_norm(s).lower() or "y") in {"y", "yes", "true", "1"}

def _int_or(s: Optional[str], default: int) -> int:
    try:
        return int(float(_norm(s)))
    except Exception:
        return default

def _header_map(ws):
    # Build case-insensitive header map: logical_name -> column_index
    headers = {}
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, name in enumerate(row1, start=1):
        if not name:
            continue
        headers[_norm(str(name)).lower()] = idx
    return headers

def _get(ws, headers, row, logical_name, *aliases):
    keys = [logical_name.lower(), *[a.lower() for a in aliases]]
    for k in keys:
        if k in headers:
            val = row[headers[k]-1]
            return _norm(str(val)) if val is not None else ""
    return ""

def _current_schedule() -> ScheduleVersion:
    sv = (ScheduleVersion.objects
          .filter(is_current=True)
          .order_by("-effective_from", "-created_at")
          .first())
    if not sv:
        raise CommandError("No current ScheduleVersion found.")
    return sv

def _resolve_vaccine(sv: ScheduleVersion, code: str, name: str) -> Optional[Vaccine]:
    code = _norm(code).lower()
    name = _norm(name)
    q = Vaccine.objects.filter(schedule_version=sv)
    if code:
        v = q.filter(code=code).first()
        if v:
            return v
    if name:
        # try exact name, then icontains as a fall-back
        v = q.filter(name__iexact=name).first()
        if v:
            return v
        return q.filter(name__icontains=name).order_by("name").first()
    return None

class Command(BaseCommand):
    help = "Import patient & doctor education videos from an .xlsx file (sheets: PE and DE)."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to .xlsx with sheets PE and DE")
        parser.add_argument("--dry-run", action="store_true", help="Parse and report, but do not write DB")

    def handle(self, *args, **opts):
        path = Path(opts["xlsx_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        wb = load_workbook(filename=str(path), data_only=True)
        if "PE" not in wb.sheetnames and "DE" not in wb.sheetnames:
            raise CommandError("Workbook must contain at least one of the sheets: PE, DE")

        sv = _current_schedule()

        created_pe = updated_pe = created_de = updated_de = 0

        with transaction.atomic():
            # --- PE sheet ---
            if "PE" in wb.sheetnames:
                ws = wb["PE"]
                headers = _header_map(ws)
                missing = [c for c in REQUIRED_PE if c not in headers]
                # vaccine_code or vaccine_name is required for matching
                if ("vaccine_code" not in headers) and ("vaccine_name" not in headers):
                    missing.append("vaccine_code or vaccine_name")
                if missing:
                    raise CommandError(f"PE: Missing columns: {missing}")

                for r in ws.iter_rows(min_row=2, values_only=True):
                    row = list(r)
                    vaccine_code = _get(ws, headers, row, "vaccine_code")
                    vaccine_name = _get(ws, headers, row, "vaccine_name", "vaccine")
                    language = _get(ws, headers, row, "language")
                    title = _get(ws, headers, row, "title")
                    video_url = _get(ws, headers, row, "video_url", "url")
                    thumb = _get(ws, headers, row, "thumbnail_url")
                    rank = _int_or(_get(ws, headers, row, "rank"), 1)
                    active = _yn(_get(ws, headers, row, "active"))

                    if not (video_url and language):
                        continue

                    vax = _resolve_vaccine(sv, vaccine_code, vaccine_name)
                    if not vax:
                        self.stdout.write(self.style.WARNING(
                            f"PE: could not match vaccine (code='{vaccine_code}', name='{vaccine_name}'), skipping"
                        ))
                        continue

                    obj, created = VaccineEducationPatient.objects.update_or_create(
                        vaccine=vax, language=language, video_url=video_url,
                        defaults={
                            "title": title,
                            "thumbnail_url": thumb,
                            "platform": VideoPlatform.YOUTUBE,
                            "rank": rank,
                            "is_active": active,
                        }
                    )
                    if created:
                        created_pe += 1
                    else:
                        updated_pe += 1

            # --- DE sheet ---
            if "DE" in wb.sheetnames:
                ws = wb["DE"]
                headers = _header_map(ws)
                missing = [c for c in REQUIRED_DE if c not in headers]
                if ("vaccine_code" not in headers) and ("vaccine_name" not in headers):
                    missing.append("vaccine_code or vaccine_name")
                if missing:
                    raise CommandError(f"DE: Missing columns: {missing}")

                for r in ws.iter_rows(min_row=2, values_only=True):
                    row = list(r)
                    vaccine_code = _get(ws, headers, row, "vaccine_code")
                    vaccine_name = _get(ws, headers, row, "vaccine_name", "vaccine")
                    title = _get(ws, headers, row, "title")
                    video_url = _get(ws, headers, row, "video_url", "url")
                    language = _get(ws, headers, row, "language") or "en"
                    rank = _int_or(_get(ws, headers, row, "rank"), 1)
                    active = _yn(_get(ws, headers, row, "active"))

                    if not video_url:
                        continue

                    vax = _resolve_vaccine(sv, vaccine_code, vaccine_name)
                    if not vax:
                        self.stdout.write(self.style.WARNING(
                            f"DE: could not match vaccine (code='{vaccine_code}', name='{vaccine_name}'), skipping"
                        ))
                        continue

                    obj, created = VaccineEducationDoctor.objects.update_or_create(
                        vaccine=vax, video_url=video_url,
                        defaults={
                            "title": title,
                            "language": language,
                            "platform": VideoPlatform.VIMEO,
                            "rank": rank,
                            "is_active": active,
                        }
                    )
                    if created:
                        created_de += 1
                    else:
                        updated_de += 1

            if opts["dry_run"]:
                self.stdout.write(self.style.WARNING("DRY RUN: rolling back changes"))
                raise CommandError("Dry-run requested; rolled back")

        self.stdout.write(self.style.SUCCESS(
            f"Imported education videos. "
            f"PE created {created_pe}, updated {updated_pe}; "
            f"DE created {created_de}, updated {updated_de}."
        ))
















